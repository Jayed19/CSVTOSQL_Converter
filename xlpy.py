from numpy import append
import openpyxl
import re

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file,read_only=True)
    sheet=workbook[sheetName]
    return sheet.max_row

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file,read_only=True)
    sheet=workbook[sheetName]
    return sheet.max_column

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file,read_only=True)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)

#XLXS File Path
filepath=r".\Test.xlsx"


#Column Finding
totalcolumn=getColumnCount(filepath,'Sheet1')
print("Total Column Count= "+str(totalcolumn))

columnnamexlxlist=[]
for column in range(1,totalcolumn+1):
    columnnamexlx=readData(filepath,'Sheet1',1,column)
    columnnamexlx=columnnamexlx.lower()
    #print("Column Name in Excell File="+columnnamexlx)
    columnnamexlxlist.append(columnnamexlx)
    
    
print(columnnamexlxlist)


#Query Template
query="UPDATE HSDL_APPLICATION SET REFERENCE_NUMBER='{ref}',LICENSE_NUMBER_EN='{dl}',APPLICATION_STATUS={app_status},CARD_STATUS={card_status},AFIS_STATUS={afis_status},ISSUE_DATE=TO_TIMESTAMP('{issue_date}','DD-MON-YYYY HH12: MI:SS:FF AM'),EXPIRY_DATE=TO_TIMESTAMP('{expiry_date}','DD-MON-YYYY HH12: MI:SS:FF AM'),STATUS={status} WHERE ID={id};"
query=query.lower()



columnsFoundInQuery=re.findall(r"\{\w+\}", query)

dict = {}

for columnnamesql in columnsFoundInQuery:
    columnnamesql=columnnamesql.replace("{","")
    columnnamesql=columnnamesql.replace("}","")
    #print("Column Name in SQL Query= "+columnnamesql)
    if columnnamesql not in columnnamexlxlist:
        print("The variable name inside {}, name "+ columnnamesql+" is not found or matched with your browsed xlxs file column name. Please check your written SQL again.")
    else:
        columnid=columnnamexlxlist.index(columnnamesql)
        dict[columnnamesql] = columnid+1

print(dict)
#print(dict.values())
    
    





#Row Finding
totalrow=getRowCount(filepath,'Sheet1')
#print("Total Row Count= "+str(totalrow))
f = open("test.sql", "a")
for row in range(2,totalrow+1):
    queryreplaced=query
    for x, y in dict.items():
        rowvalue=(readData(filepath,'Sheet1',row,y))
        
        x="{"+x+"}"
        #print(x + " = "+str(rowvalue))
        
        
        queryreplaced=queryreplaced.replace(x,str(rowvalue))
    #print(queryreplaced)
    
    f.write(queryreplaced)
    f.write("\n")

f.close()